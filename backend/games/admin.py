"""
Admin configuration for games app with XLSX export functionality.
"""
from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import (
    UserProfile,
    GameSession,
    Leaderboard,
    Achievement,
    UserAchievement,
    Friendship
)


def export_game_sessions_to_xlsx(modeladmin, request, queryset):
    """
    Экспорт выбранных игровых сессий в XLSX файл.
    """
    # Создаем workbook и активный лист
    wb = Workbook()
    ws = wb.active
    ws.title = "Игровые сессии"
    
    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Заголовки столбцов
    headers = [
        'ID', 'Пользователь', 'Email', 'Очки', 'Сложность', 
        'Время игры (сек)', 'Завершена', 'Средняя реакция (мс)', 
        'Дата создания'
    ]
    
    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Записываем данные
    for row_num, session in enumerate(queryset.select_related('user'), 2):
        ws.cell(row=row_num, column=1, value=session.id)
        ws.cell(row=row_num, column=2, value=session.user.username)
        ws.cell(row=row_num, column=3, value=session.user.email)
        ws.cell(row=row_num, column=4, value=session.score)
        ws.cell(row=row_num, column=5, value=session.get_difficulty_display())
        ws.cell(row=row_num, column=6, value=session.time_played)
        ws.cell(row=row_num, column=7, value='Да' if session.is_completed else 'Нет')
        ws.cell(row=row_num, column=8, value=round(session.avg_reaction_time, 2) if session.avg_reaction_time else '-')
        ws.cell(row=row_num, column=9, value=session.created_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    # Автоматическая ширина столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Создаем HTTP ответ с файлом
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'game_sessions_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


export_game_sessions_to_xlsx.short_description = "Экспортировать выбранные сессии в XLSX"


def export_leaderboard_to_xlsx(modeladmin, request, queryset):
    """
    Экспорт таблицы лидеров в XLSX файл.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица лидеров"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Заголовки
    headers = [
        'Ранг', 'Пользователь', 'Email', 'Очки', 'Сложность',
        'Средняя реакция (мс)', 'Дата достижения'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Данные
    for row_num, entry in enumerate(queryset.select_related('user').order_by('rank'), 2):
        ws.cell(row=row_num, column=1, value=entry.rank or row_num - 1)
        ws.cell(row=row_num, column=2, value=entry.user.username)
        ws.cell(row=row_num, column=3, value=entry.user.email)
        ws.cell(row=row_num, column=4, value=entry.score)
        ws.cell(row=row_num, column=5, value=entry.get_difficulty_display())
        ws.cell(row=row_num, column=6, value=round(entry.avg_reaction_time, 2) if entry.avg_reaction_time else '-')
        ws.cell(row=row_num, column=7, value=entry.date_achieved.strftime('%Y-%m-%d %H:%M:%S'))
    
    # Автоширина
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'leaderboard_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


export_leaderboard_to_xlsx.short_description = "Экспортировать в XLSX"


@admin.register(UserProfile)
class UserProfileAdmin(admin.ModelAdmin):
    list_display = ('user', 'bio', 'date_of_birth', 'created_at')
    search_fields = ('user__username', 'user__email', 'bio')
    list_filter = ('created_at',)
    readonly_fields = ('created_at', 'updated_at')


@admin.register(GameSession)
class GameSessionAdmin(admin.ModelAdmin):
    list_display = ('id', 'user', 'score', 'difficulty', 'is_completed', 'avg_reaction_time', 'created_at')
    list_filter = ('difficulty', 'is_completed', 'created_at')
    search_fields = ('user__username', 'user__email')
    readonly_fields = ('created_at', 'updated_at', 'avg_reaction_time')
    actions = [export_game_sessions_to_xlsx]
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('user', 'score', 'difficulty', 'is_completed')
        }),
        ('Игровые данные', {
            'fields': ('game_state', 'time_played', 'reaction_times', 'avg_reaction_time')
        }),
        ('Временные метки', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )


@admin.register(Leaderboard)
class LeaderboardAdmin(admin.ModelAdmin):
    list_display = ('rank', 'user', 'score', 'difficulty', 'avg_reaction_time', 'date_achieved')
    list_filter = ('difficulty', 'date_achieved')
    search_fields = ('user__username', 'user__email')
    readonly_fields = ('created_at', 'updated_at')
    actions = [export_leaderboard_to_xlsx]
    ordering = ['rank', '-score']


@admin.register(Achievement)
class AchievementAdmin(admin.ModelAdmin):
    list_display = ('name', 'achievement_type', 'points', 'created_at')
    list_filter = ('achievement_type', 'created_at')
    search_fields = ('name', 'description')
    readonly_fields = ('created_at', 'updated_at')


@admin.register(UserAchievement)
class UserAchievementAdmin(admin.ModelAdmin):
    list_display = ('user', 'achievement', 'unlocked_at')
    list_filter = ('unlocked_at', 'achievement')
    search_fields = ('user__username', 'achievement__name')
    readonly_fields = ('created_at', 'updated_at', 'unlocked_at')
    
    def has_add_permission(self, request):
        """Администратор может вручную назначать достижения."""
        return True


@admin.register(Friendship)
class FriendshipAdmin(admin.ModelAdmin):
    list_display = ('from_user', 'to_user', 'status', 'created_at')
    list_filter = ('status', 'created_at')
    search_fields = ('from_user__username', 'to_user__username')
    readonly_fields = ('created_at', 'updated_at')
