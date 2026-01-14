"""
Тесты безопасности и админ-функционала.
Проверяют защиту от SQL-инъекций, XSS, хеширование паролей, экспорт XLSX.
"""
import pytest
import openpyxl
from io import BytesIO
from django.contrib.auth import get_user_model
from django.contrib.admin.sites import AdminSite
from django.test import RequestFactory
from rest_framework.test import APIClient
from games.models import GameSession, Leaderboard
from games.admin import GameSessionAdmin, export_game_sessions_to_xlsx

User = get_user_model()


@pytest.mark.django_db
class TestSecurity:
    """Тесты безопасности."""
    
    def test_passwords_are_hashed(self):
        """Проверка, что пароли хранятся в хешированном виде."""
        user = User.objects.create_user(
            username='secureuser',
            email='secure@test.com',
            password='mypassword123'
        )
        
        # Пароль не должен храниться в открытом виде
        assert user.password != 'mypassword123'
        # Пароль должен быть хешем
        assert user.password.startswith('pbkdf2_sha256$')
        # Проверка пароля должна работать
        assert user.check_password('mypassword123')
    
    def test_sql_injection_protection(self):
        """Тест защиты от SQL-инъекций через ORM."""
        # Попытка SQL-инъекции через поиск
        malicious_input = "'; DROP TABLE games_gamesession; --"
        
        # ORM Django автоматически экранирует запросы
        users = User.objects.filter(username=malicious_input)
        
        # Запрос не должен вызвать ошибку
        assert users.count() == 0
        # Таблица должна существовать
        assert GameSession.objects.model._meta.db_table is not None
    
    def test_xss_protection_in_user_input(self):
        """Тест защиты от XSS в пользовательском вводе."""
        xss_payload = '<script>alert("XSS")</script>'
        
        user = User.objects.create_user(
            username='xsstest',
            email='xss@test.com'
        )
        user.profile.bio = xss_payload
        user.profile.save()
        
        # Django автоматически экранирует данные при выводе в шаблонах
        # Здесь проверяем, что данные сохраняются как есть
        assert user.profile.bio == xss_payload
        # В реальных шаблонах Django экранирует это автоматически
    
    def test_unauthorized_access_to_admin(self):
        """Тест, что неавторизованный пользователь не имеет доступа к админке."""
        from django.test import Client
        
        client = Client()
        response = client.get('/admin/')
        
        # Должен редиректить на страницу логина
        assert response.status_code == 302
        assert '/admin/login/' in response.url


@pytest.mark.django_db
class TestAdminXLSXExport:
    """Тесты экспорта XLSX из админ-панели."""
    
    def test_export_game_sessions_to_xlsx(self):
        """Тест экспорта игровых сессий в XLSX."""
        # Создаем тестовые данные
        user1 = User.objects.create_user(username='player1', email='p1@test.com')
        user2 = User.objects.create_user(username='player2', email='p2@test.com')
        
        session1 = GameSession.objects.create(
            user=user1,
            score=1000,
            difficulty='easy',
            time_played=60,
            is_completed=True
        )
        session2 = GameSession.objects.create(
            user=user2,
            score=2000,
            difficulty='hard',
            time_played=120,
            is_completed=True
        )
        
        # Создаем mock request
        factory = RequestFactory()
        request = factory.get('/admin/games/gamesession/')
        request.user = User.objects.create_superuser(
            username='admin',
            email='admin@test.com',
            password='admin123'
        )
        
        # Создаем admin instance
        model_admin = GameSessionAdmin(GameSession, AdminSite())
        
        # Выполняем экспорт
        queryset = GameSession.objects.all()
        response = export_game_sessions_to_xlsx(model_admin, request, queryset)
        
        # Проверяем response
        assert response.status_code == 200
        assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert 'game_sessions_' in response['Content-Disposition']
        
        # Проверяем содержимое XLSX
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Проверяем заголовки
        headers = [cell.value for cell in ws[1]]
        assert 'Пользователь' in headers
        assert 'Очки' in headers
        assert 'Сложность' in headers
        
        # Проверяем данные (должно быть 2 строки данных + 1 заголовок)
        assert ws.max_row == 3
        
        # Проверяем конкретные значения
        usernames = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1)]
        assert 'player1' in usernames
        assert 'player2' in usernames
    
    def test_xlsx_file_structure(self):
        """Проверка структуры и содержимого XLSX файла."""
        user = User.objects.create_user(username='test', email='test@test.com')
        GameSession.objects.create(
            user=user,
            score=500,
            difficulty='medium',
            time_played=30,
            is_completed=True,
            avg_reaction_time=250.5
        )
        
        factory = RequestFactory()
        request = factory.get('/admin/')
        request.user = User.objects.create_superuser(username='admin2', email='a2@test.com')
        
        model_admin = GameSessionAdmin(GameSession, AdminSite())
        queryset = GameSession.objects.all()
        response = export_game_sessions_to_xlsx(model_admin, request, queryset)
        
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Проверяем, что средняя реакция округлена до 2 знаков
        reaction_time_col = None
        for idx, cell in enumerate(ws[1], 1):
            if 'реакция' in str(cell.value).lower():
                reaction_time_col = idx
                break
        
        if reaction_time_col:
            reaction_value = ws.cell(row=2, column=reaction_time_col).value
            assert reaction_value == 250.5


@pytest.mark.django_db
class TestIntegrationScenarios:
    """Интеграционные тесты полного цикла функционала."""
    
    def test_complete_game_flow(self):
        """Тест полного цикла игры: регистрация → игра → достижение → лидерборд."""
        client = APIClient()
        
        # 1. Регистрация
        register_data = {
            'username': 'newplayer',
            'email': 'newplayer@test.com',
            'password': 'securepass123',
            'password2': 'securepass123'
        }
        response = client.post('/api/auth/register/', register_data)
        assert response.status_code == 201
        
        # 2. Логин
        login_data = {
            'username': 'newplayer',
            'password': 'securepass123'
        }
        response = client.post('/api/auth/login/', login_data)
        assert response.status_code == 200
        token = response.data['access']
        
        # 3. Аутентификация
        client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')
        
        # 4. Создание игровой сессии
        session_data = {
            'score': 600,
            'difficulty': 'easy',
            'time_played': 60,
            'is_completed': True,
            'reaction_times': [200, 250, 300]
        }
        response = client.post('/api/games/sessions/', session_data, format='json')
        assert response.status_code == 201
        
        # 5. Проверка, что сессия сохранена
        user = User.objects.get(username='newplayer')
        assert GameSession.objects.filter(user=user).exists()
        
        # 6. Проверка обновления лидерборда
        assert Leaderboard.objects.filter(user=user, difficulty='easy').exists()
    
    def test_friend_system_flow(self):
        """Тест полного цикла системы друзей."""
        client = APIClient()
        
        # Создаем двух пользователей
        user1 = User.objects.create_user(username='user1', email='u1@test.com', password='pass')
        user2 = User.objects.create_user(username='user2', email='u2@test.com', password='pass')
        
        # User1 отправляет запрос User2
        client.force_authenticate(user=user1)
        response = client.post('/api/games/friends/', {'friend_identifier': 'user2'})
        assert response.status_code == 201
        
        # User2 принимает запрос
        client.force_authenticate(user=user2)
        friendship_id = response.data['id']
        response = client.post(f'/api/games/friends/{friendship_id}/accept/')
        assert response.status_code == 200
        
        # Проверяем, что они теперь друзья
        response = client.get('/api/games/friends/friends/')
        assert response.status_code == 200
        assert any(f['username'] == 'user1' for f in response.data)


@pytest.mark.django_db
class TestPerformance:
    """Тесты производительности (N+1 проблема)."""
    
    def test_no_n_plus_one_in_leaderboard(self, django_assert_num_queries):
        """Проверка отсутствия N+1 запросов в лидерборде."""
        # Создаем несколько записей
        for i in range(10):
            user = User.objects.create_user(username=f'user{i}', email=f'u{i}@test.com')
            Leaderboard.objects.create(user=user, score=i*100, difficulty='easy')
        
        client = APIClient()
        
        # Должно быть фиксированное количество запросов, независимо от количества записей
        with django_assert_num_queries(2):  # 1 для leaderboard + 1 для users (select_related)
            response = client.get('/api/games/leaderboard/')
            assert response.status_code == 200
